import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import numbers

def do_a_summary(my_dir, multiplier=1.0):
    # List all excel files in current folder
    template_name = 'summary_template.xlsx'
    summary_name = 'summary.xlsx'
    cur_filelist = [f for f in os.listdir(my_dir)
                    if os.path.isfile(os.path.join(my_dir, f))
                    and (f.endswith(".xls") or f.endswith(".xlsx"))]
    if template_name not in cur_filelist:
        return ("No template file found in {}. Please make sure a template "
                "file exists with the name {}, or select a different "
                "directory.").format(my_dir, template_name)
    cur_filelist.remove(template_name)
    # Scan template file to find indicators
    try:
        wb_t_path = os.path.join(my_dir, template_name)
        wb_t = load_workbook(wb_t_path)
    except Exception as e:
        return ("Unable to load template file {}, the following error message "
                "was returned:-\n{}").format(wb_t_path, str(e))
    locations = {}
    for sheet_name in wb_t.sheetnames:
        sheet = wb_t[sheet_name]
        for row in sheet:
            for c in row:
                if isinstance(c.value, str):
                    if c.value.startswith('::') and c.value.endswith('::'):
                        row_identifier = c.value[2:-2]
                        locations[row_identifier] = {'sheet_name': sheet_name,
                                                    'col': c.column,
                                                    'row': c.row}
    if len(locations) == 0:
        return ("No markers found in template file {}, please make sure "
                "it contains at least one cell with the value '::?::' "
                "where '?' is a valid Excel row name (alphabets A-Z, AA-ZZ "
                "etc.)").format(wb_t_path)
    # Create a summary file if it doesn't exist, otherwise load the current
    # one and append to it
    if summary_name in cur_filelist:
        # Load it up, then remove from cur_filelist
        wb_s_path = os.path.join(my_dir, summary_name)
        try:
            wb_s = load_workbook(wb_s_path)
        except Exception as e:
            return ("Unable to load summary file {}, the following error "
                    "message was returned:-\n{}").format(wb_s_path, str(e))
        ws_s = wb_s.active
        cur_filelist.remove(summary_name)
    else:
        wb_s = Workbook()
        ws_s = wb_s.active
        ws_s.title = "Summary of data"
    # Process all other excel files, log all errors appropriately
    error_list = []
    summary_list = []
    for f in cur_filelist:
        wb_path = os.path.join(my_dir, f)
        try:
            # Open file
            wb = load_workbook(wb_path, data_only=True)
            summary_row = {}
            for summary_col, identifier in locations.items():
                # Extract values at locations indicated in template file
                sheet = wb[identifier['sheet_name']]
                col = identifier['col']
                row = identifier['row']
                val = sheet.cell(row, col).value
                if isinstance(val, numbers.Number) and val < 100*multiplier:
                    summary_row[summary_col] = val*multiplier
                else:
                    summary_row[summary_col] = val
            summary_list.append(summary_row)
        except Exception as e:
            error_list.append((wb_path, e))
    def filter_part_one(summary_row):
        count_of_zeros = 0
        for col, val in summary_row.items():
            if val == 0:
                count_of_zeros += 1
        if count_of_zeros > 12:
            retval = {}
            for col, val in summary_row.items():
                if isinstance(val, numbers.Number) and val < 100*multiplier:
                    retval[col] = 'P'
                else:
                    retval[col] = val
        else:
            retval = summary_row
        return retval
    filtered_list = [filter_part_one(r) for r in summary_list]
    # Create sorted list based on first column
    sorted_list = sorted(filtered_list, key=lambda r: r['A'])
    # Actually write the rows to the openpyxl object
    for summary_row in sorted_list:
        # Create new row in summary file (check for old?)
        row = ws_s.max_row + 1
        for col, val in summary_row.items():
            coordinate = col + str(row)
            ws_s.cell(column=column_index_from_string(col), row=row, value=val)
    # Save the file
    try:
        wb_s.save(filename=os.path.join(my_dir, summary_name))
    except Exception as e:
        return ("Unable to save summary file {}, the following error message "
                "was returned:-\n{}").format(wb_s-path, str(e))
    if len(error_list):  # Some error(s) occurred, report them
        errors = "\n".join(["Error for file {} is '{}'".format(p, str(e))
                            for (p, e) in error_list])
        return ("Errors found\n"
                "{}\n"
                "All other files successfully summarized.").format(errors)
    else:
        return ("Successful summary done, please choose another directory "
                "if necessary.")
